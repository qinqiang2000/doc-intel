"""S4: Excel rendering for an evaluation run."""
from __future__ import annotations

import io
from collections import defaultdict
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.models.evaluation_field_result import EvaluationFieldResult
from app.models.evaluation_run import EvaluationRun


_FILL = {
    "exact": PatternFill(fill_type="solid", fgColor="C6EFCE"),
    "fuzzy": PatternFill(fill_type="solid", fgColor="FFEB9C"),
    "mismatch": PatternFill(fill_type="solid", fgColor="FFC7CE"),
    "missing_pred": PatternFill(fill_type="solid", fgColor="D9D9D9"),
    "missing_expected": PatternFill(fill_type="none"),
}


def render_run_xlsx(run: EvaluationRun, fields: Iterable[EvaluationFieldResult]) -> bytes:
    fields = list(fields)
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    detail = wb.create_sheet("Detail")

    bold = Font(bold=True)

    # ===== Summary sheet =====
    summary.append([
        "field_name", "exact", "fuzzy", "mismatch",
        "missing_pred", "missing_expected", "accuracy",
    ])
    for cell in summary[1]:
        cell.font = bold

    counts: dict[str, dict[str, int]] = defaultdict(lambda: {
        "exact": 0, "fuzzy": 0, "mismatch": 0,
        "missing_pred": 0, "missing_expected": 0,
    })
    for fr in fields:
        counts[fr.field_name][fr.match_status] += 1

    total = {"exact": 0, "fuzzy": 0, "mismatch": 0, "missing_pred": 0, "missing_expected": 0}
    for field_name in sorted(counts):
        c = counts[field_name]
        denom = c["exact"] + c["fuzzy"] + c["mismatch"] + c["missing_pred"]
        accuracy = (c["exact"] + c["fuzzy"]) / denom if denom else 0.0
        summary.append([
            field_name, c["exact"], c["fuzzy"], c["mismatch"],
            c["missing_pred"], c["missing_expected"], round(accuracy, 4),
        ])
        for k in total:
            total[k] += c[k]

    total_denom = total["exact"] + total["fuzzy"] + total["mismatch"] + total["missing_pred"]
    total_acc = (total["exact"] + total["fuzzy"]) / total_denom if total_denom else 0.0
    total_row = summary.max_row + 1
    summary.append([
        "TOTAL", total["exact"], total["fuzzy"], total["mismatch"],
        total["missing_pred"], total["missing_expected"], round(total_acc, 4),
    ])
    for cell in summary[total_row]:
        cell.font = bold

    # Column widths
    summary.column_dimensions["A"].width = 30
    for col in "BCDEFG":
        summary.column_dimensions[col].width = 14

    # ===== Detail sheet =====
    detail.append(["filename", "field_name", "predicted", "expected", "status"])
    for cell in detail[1]:
        cell.font = bold
    for fr in sorted(fields, key=lambda x: (x.document_filename, x.field_name)):
        detail.append([
            fr.document_filename, fr.field_name,
            fr.predicted_value or "",
            fr.expected_value or "",
            fr.match_status,
        ])
        status_cell = detail.cell(row=detail.max_row, column=5)
        fill = _FILL.get(fr.match_status)
        if fill is not None:
            status_cell.fill = fill

    detail.column_dimensions["A"].width = 30
    detail.column_dimensions["B"].width = 30
    detail.column_dimensions["C"].width = 40
    detail.column_dimensions["D"].width = 40
    detail.column_dimensions["E"].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
